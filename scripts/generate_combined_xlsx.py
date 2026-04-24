"""Combine all 13 pavement sections into a single .xlsx with one sheet per section.

Each sheet contains: section metadata, design parameters (R=360, SCI=80, growth,
life), layer structure, top-10 aircraft mix, and a cumulative CDF column so the
recipient can see how much of the section CDF each aircraft contributes.

Output: results/CEE598_AllSections_For_Desktop_Calibration.xlsx
"""
import json
import os
import sys

PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(PROJECT, 'scripts'))

# Pull the AIRPORTS dict + history from the batch script (for layer structures).
import importlib.util, ast
spec_path = os.path.join(PROJECT, 'scripts', 'all_airports_cdf_with_sci_history.py')
src = open(spec_path).read()
tree = ast.parse(src)
keep_names = ('mor_for_pcc_age',)
keep_assigns = ('FLEX_STR','AGED_MOR','AGED_THRESHOLD_YRS','DESIGN_LIFE','MIN_MTOW',
                'CURRENT_YEAR','HISTORY','AIRPORTS')
keep = ast.Module(body=[
    n for n in tree.body
    if (isinstance(n, ast.FunctionDef) and n.name in keep_names)
       or (isinstance(n, ast.Assign)
           and any(isinstance(t, ast.Name) and t.id in keep_assigns for t in n.targets))
], type_ignores=[])
ns = {'__name__': '__import__', '__file__': spec_path}
exec(compile(keep, spec_path, 'exec'), ns)
AIRPORTS = ns['AIRPORTS']
HISTORY  = ns['HISTORY']
mor_for_pcc_age = ns['mor_for_pcc_age']

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

CDF_RESULTS = os.path.join(PROJECT, 'results', 'cdf_results.json')
GROUPS_DIR = os.path.join(PROJECT, 'results', 'aircraft_groups')
OUT_PATH = os.path.join(PROJECT, 'results', 'CEE598_AllSections_For_Desktop_Calibration.xlsx')

# Styling
THIN = Side(border_style='thin', color='808080')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
HDR_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SECT_FILL = PatternFill('solid', fgColor='B4C7E7')
SECT_FONT = Font(name='Calibri', size=11, bold=True, color='000000')
KEY_FONT = Font(name='Calibri', size=10, bold=True)
VAL_FONT = Font(name='Calibri', size=10)
NOTE_FONT = Font(name='Calibri', size=9, italic=True, color='707070')
UNDER_FILL = PatternFill('solid', fgColor='F4CCCC')
OVER_FILL = PatternFill('solid', fgColor='D9EAD3')


def load_top10_csv(icao, sid):
    csv_path = os.path.join(GROUPS_DIR, f'{icao}_{sid}.csv')
    if not os.path.exists(csv_path):
        return []
    rows = []
    with open(csv_path) as f:
        next(f)  # header
        for line in f:
            parts = line.strip().split(',')
            if len(parts) < 8: continue
            rows.append({
                'rank': int(parts[0]),
                'icao': parts[1],
                'mtow': float(parts[2]),
                'gear': parts[3],
                'annual_deps': float(parts[4]),
                'growth_rate': float(parts[5]),
                'cdf_contribution': float(parts[6]),
                'source': parts[7],
            })
    return rows


def fill_index_sheet(ws, results):
    ws.title = 'Index'
    ws['A1'] = 'CEE 598 Final Project — All Pavement Sections'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E78')
    ws.merge_cells('A1:H1')
    ws['A2'] = 'For desktop FAARFIELD 2.1.1 calibration. Each sheet = one pavement section.'
    ws['A2'].font = NOTE_FONT
    ws.merge_cells('A2:H2')

    ws['A4'] = 'Methodology summary:'
    ws['A4'].font = KEY_FONT
    rows = [
        ('Engine', 'FAARFIELD desktop-parity native backend (LEAFClassLib + AMClassLib FAASR3D 3D rigid FEM)'),
        ('PCC R (MOR)', '360 psi (FAA AC 150/5320-6F lower bound: 0.08 × f\'c, f\'c = 4500 psi assumed for >20-yr aged PCC)'),
        ('PCC SCI', '80 (FAARFIELD design default — moderate / first observable cracking)'),
        ('Design life', '20 years'),
        ('Wander σ', '30.435 in (FAA standard for rigid pavement, 41-offset lateral sweep 0–400")'),
        ('Verdict overall', f'{sum(1 for r in results if r["adequate"])} OVER / {sum(1 for r in results if not r["adequate"])} UNDER (CDF<1.0 = OVER)'),
        ('Generated', '2026-04-22 (R=360 psi / SCI=80 batch)'),
    ]
    for i, (k, v) in enumerate(rows, 5):
        ws.cell(row=i, column=1, value=k).font = KEY_FONT
        ws.cell(row=i, column=2, value=v).font = VAL_FONT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

    # Section list table
    start = 5 + len(rows) + 2
    headers = ['Sheet', 'Airport', 'Section', 'Use', 'Structure', 'CDF_max', 'Verdict', 'Controlling']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    for i, r in enumerate(sorted(results, key=lambda x: (x['icao'], x['section_id']))):
        row = start + 1 + i
        sheet_name = f"{r['icao']}_{r['section_id']}"
        ws.cell(row=row, column=1, value=sheet_name).hyperlink = f"#{sheet_name}!A1"
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10, color='0563C1', underline='single')
        ws.cell(row=row, column=2, value=r['airport'])
        ws.cell(row=row, column=3, value=r['section_id'])
        ws.cell(row=row, column=4, value=r['use'])
        ws.cell(row=row, column=5, value=r['desc'])
        cdf_cell = ws.cell(row=row, column=6, value=r['cdf_max'])
        cdf_cell.number_format = '0.000E+00'
        verdict = 'OVER' if r['adequate'] else 'UNDER'
        v_cell = ws.cell(row=row, column=7, value=verdict)
        v_cell.fill = OVER_FILL if r['adequate'] else UNDER_FILL
        v_cell.font = Font(name='Calibri', size=10, bold=True)
        v_cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=8, value=r['controlling'])
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = BORDER

    for c, w in enumerate([14, 28, 10, 16, 28, 13, 9, 18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def fill_section_sheet(ws, rec, sec_struct, top10):
    ws.title = f"{rec['icao']}_{rec['section_id']}"
    icao = rec['icao']
    sid = rec['section_id']
    growth = rec.get('growth', 0)
    mor = rec.get('mor_psi', 360)
    sci = rec.get('sci', 80)
    pcc_year = rec.get('pcc_year')
    pcc_age = rec.get('pcc_age_yrs')

    # ============== HEADER ==============
    ws['A1'] = f"{icao} — Section {sid} — {rec['use']}"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E78')
    ws.merge_cells('A1:F1')

    ws['A2'] = f"{rec['airport']}  |  Structure: {rec['desc']}"
    ws['A2'].font = NOTE_FONT
    ws.merge_cells('A2:F2')

    verdict = 'OVER' if rec['adequate'] else 'UNDER'
    ws['A3'] = f"Native-engine CDF_max = {rec['cdf_max']:.3e}  ({verdict})  |  Controlling: {rec['controlling']}"
    ws['A3'].font = Font(name='Calibri', size=11, bold=True,
                          color='8B0000' if not rec['adequate'] else '006100')
    ws.merge_cells('A3:F3')

    # ============== DESIGN PARAMETERS ==============
    row = 5
    ws.cell(row=row, column=1, value='Design Parameters').font = SECT_FONT
    ws.cell(row=row, column=1).fill = SECT_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    params = [
        ('Design type', 'Flex on Rigid (HMA Overlay over Existing Rigid)'),
        ('Design life (yr)', 20),
        ('Growth rate (per year, decimal)', growth),
        ('Growth rate (% per year)', f"{growth*100:+.3f} %"),
        ('PCC flexural strength R (psi)', mor),
        ('PCC SCI', sci),
        ('PCC pour year', pcc_year if pcc_year else '?'),
        ('PCC age at analysis (yr)', pcc_age if pcc_age else '?'),
        ('Use 3D FEM (rigid pavement)', 'Yes (AMClassLib FAASR3D — desktop default for HMA-on-rigid)'),
        ('Wander σ (in)', '30.435 (FAA standard for rigid pavement)'),
        ('Lateral sweep', '41 offsets, 0 to 400 in at 10 in increments'),
        ('R rule source', 'FAA AC 150/5320-6F: R = 0.08 × f\'c, f\'c = 4500 psi assumed for >20-yr aged PCC'),
    ]
    for k, v in params:
        ws.cell(row=row, column=1, value=k).font = KEY_FONT
        ws.cell(row=row, column=2, value=v).font = VAL_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    row += 1

    # ============== LAYER STRUCTURE ==============
    ws.cell(row=row, column=1, value='Layer Structure (top → bottom)').font = SECT_FONT
    ws.cell(row=row, column=1).fill = SECT_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    layer_headers = ['#', 'Layer Type', 'Thickness (in)', 'E (psi)', 'Poisson ν', 'Notes']
    for c, h in enumerate(layer_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1
    layers = sec_struct['layers']
    for i, L in enumerate(layers, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=L['type'])
        ws.cell(row=row, column=3, value=L['h'])
        ws.cell(row=row, column=4, value=L['E']).number_format = '#,##0'
        ws.cell(row=row, column=5, value=L['nu'])
        notes = ''
        if 'PCC' in L['type']:
            notes = f'Override flexural strength to R = {mor} psi'
        ws.cell(row=row, column=6, value=notes).font = NOTE_FONT
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER
        row += 1
    sg = sec_struct['subgrade']
    ws.cell(row=row, column=1, value=len(layers)+1)
    sub_desc = next((a['subgrade_desc'] for a in AIRPORTS if a['icao'] == icao), '?')
    ws.cell(row=row, column=2, value=f'Subgrade — {sub_desc}')
    ws.cell(row=row, column=3, value='infinite')
    ws.cell(row=row, column=4, value=sg['E']).number_format = '#,##0'
    ws.cell(row=row, column=5, value=sg['nu'])
    ws.cell(row=row, column=6, value='').font = NOTE_FONT
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = BORDER
    row += 2

    # ============== AIRCRAFT MIX with CUMULATIVE CDF ==============
    ws.cell(row=row, column=1, value='Aircraft Mix (Top 10 by CDF Contribution)').font = SECT_FONT
    ws.cell(row=row, column=1).fill = SECT_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    notes = ws.cell(row=row, column=1, value='Ranks 1-5: FEM-augmented (batch).  Ranks 6-10: LEAF-only supplement (CDF magnitudes approximate, ranking preserved).')
    notes.font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    ac_headers = ['Rank', 'ICAO', 'MTOW (lb)', 'Gear', 'Annual Departures',
                  'Growth Rate', 'CDF Contribution', 'Cumulative CDF', 'Source']
    for c, h in enumerate(ac_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    row += 1

    cum = 0.0
    section_total = sum(a['cdf_contribution'] for a in top10) if top10 else 0
    for a in top10:
        cum += a['cdf_contribution']
        ws.cell(row=row, column=1, value=a['rank'])
        ws.cell(row=row, column=2, value=a['icao'])
        ws.cell(row=row, column=3, value=a['mtow']).number_format = '#,##0'
        ws.cell(row=row, column=4, value=a['gear'])
        ws.cell(row=row, column=5, value=a['annual_deps']).number_format = '0.00'
        ws.cell(row=row, column=6, value=a['growth_rate']).number_format = '0.0000'
        ws.cell(row=row, column=7, value=a['cdf_contribution']).number_format = '0.000E+00'
        ws.cell(row=row, column=8, value=cum).number_format = '0.000E+00'
        ws.cell(row=row, column=9, value=a['source']).font = NOTE_FONT
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = BORDER
        row += 1

    # Cumulative CDF total for top 10 + comparison to native section CDF
    row += 1
    ws.cell(row=row, column=1, value='Cumulative top-10 CDF:').font = KEY_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=8, value=cum).number_format = '0.000E+00'
    ws.cell(row=row, column=8).font = Font(name='Calibri', size=11, bold=True)
    row += 1
    ws.cell(row=row, column=1, value='Native section CDF_max (from 41-offset sweep, all aircraft, FEM-augmented):').font = KEY_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    nv_cell = ws.cell(row=row, column=8, value=rec['cdf_max'])
    nv_cell.number_format = '0.000E+00'
    nv_cell.font = Font(name='Calibri', size=11, bold=True,
                        color='8B0000' if not rec['adequate'] else '006100')
    row += 2

    # CDF breakdown by failure mode
    ws.cell(row=row, column=1, value='CDF Breakdown by Failure Mode').font = SECT_FONT
    ws.cell(row=row, column=1).fill = SECT_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    breakdown = [
        ('AC fatigue (RDEC)', rec['cdf_ac']),
        ('Subgrade rutting', rec['cdf_sub']),
        ('PCC fatigue (LeafCDFRigid_2014)', rec['cdf_pcc']),
        ('Section CDF_max (controlling)', rec['cdf_max']),
    ]
    for k, v in breakdown:
        ws.cell(row=row, column=1, value=k).font = KEY_FONT
        cell = ws.cell(row=row, column=2, value=v)
        cell.number_format = '0.000E+00'
        if k.startswith('Section'):
            cell.font = Font(name='Calibri', size=11, bold=True,
                             color='8B0000' if not rec['adequate'] else '006100')
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER
        row += 1

    # Column widths
    widths = [10, 26, 15, 8, 17, 14, 18, 18, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def main():
    print(f"Building {OUT_PATH}")
    saved = json.load(open(CDF_RESULTS))
    saved_by_key = {(r['icao'], r['section_id']): r for r in saved}

    wb = openpyxl.Workbook()

    # Index sheet first
    idx_ws = wb.active
    fill_index_sheet(idx_ws, saved)

    # One sheet per section
    for apt in AIRPORTS:
        for sec in apt['sections']:
            rec = saved_by_key.get((apt['icao'], sec['id']))
            if not rec:
                print(f"  {apt['icao']} {sec['id']}: NOT in cdf_results - skip")
                continue
            top10 = load_top10_csv(apt['icao'], sec['id'])
            ws = wb.create_sheet(title=f"{apt['icao']}_{sec['id']}")
            fill_section_sheet(ws, rec, sec, top10)
            print(f"  {apt['icao']} {sec['id']}: sheet written ({len(top10)} aircraft)")

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")


if __name__ == '__main__':
    main()
